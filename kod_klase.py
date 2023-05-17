import psycopg2 as pg
import openpyxl as op
import pandas as pd

class Maratonac:
    def __init__(self,rednibroj,imeprezime,godiste,zemlja,kategorija):
        self.rednibroj=rednibroj
        self.imeprezime=imeprezime
        self.godiste=godiste
        self.zemlja=zemlja
        self.kategorija=kategorija
    
    def __str__(self):
        return "{} {} {}".format(self.rednibroj,self.imeprezime,self.zemlja)
    

class Maratonci:
    def __init__(self):
        self.l_m=pd.Series()
    
    def ucitaj_maratonce(self):
        try:
            con=pg.connect(database='Maraton',
                           user='postgres',
                           host='localhost',
                           port='5432',
                           password='itoip')
            
            cursor=con.cursor()
            cursor.execute('SELECT * FROM MARATONCI')

            r=pd.Series(cursor.fetchall())

            for i in r:
                m=Maratonac(i[0],i[1],i[2],i[3],i[4])
                self.l_m[len(self.l_m)]=m
        except (Exception,pg.Error) as e:
            return "Error",e
        finally:
            cursor.close()
            con.close()

    def search_listbox(self,imeprezime):
        l=[]
        for i in self.l_m:
            if imeprezime in i.imeprezime.lower():
                l.append(i)
        return l
    
    def export_maratonac(self,maratonac):
        wb=op.Workbook()
        ws=wb.active
        ws.title=maratonac.imeprezime

        ws['A1']="Redni Broj:"
        ws['A2']="Ime prezime:"
        ws['A3']="Godiste:"
        ws['A4']="Zemlja:"
        ws['A5']="Kategorija:"

        ws['B1']=maratonac.rednibroj
        ws['B2']=maratonac.imeprezime
        ws['B3']=maratonac.godiste
        ws['B4']=maratonac.zemlja
        ws['B5']=maratonac.kategorija

        wb.save(filename=maratonac.imeprezime+'.xlsx')

    def export_sve_excel(self):
        wb=op.Workbook()
        ws=wb.active
        ws.title='Svi maratonci'

        ws['A1']="Redni Broj:"
        ws['B1']="Ime prezime:"
        ws['C1']="Godiste:"
        ws['D1']="Zemlja:"
        ws['E1']="Kategorija:"
        k=2

        for i in self.l_m:
            ws.cell(row=k,column=1).value=i.rednibroj
            ws.cell(row=k,column=2).value=i.imeprezime
            ws.cell(row=k,column=3).value=i.godiste
            ws.cell(row=k,column=4).value=i.zemlja
            ws.cell(row=k,column=5).value=i.kategorija
            k+=1

        wb.save(filename='sve.xlsx')    

    def export_ime_zemlja_excel(self):
        wb=op.Workbook()
        ws=wb.active
        ws.title='Ime zemljamaratonci'

        ws['A1']="Redni Broj:"
        ws['B1']="Ime prezime:"
        ws['C1']="Zemlja:"
        k=2

        for i in self.l_m:
            ws.cell(row=k,column=1).value=i.rednibroj
            ws.cell(row=k,column=2).value=i.imeprezime
            ws.cell(row=k,column=3).value=i.zemlja
            k+=1

        wb.save(filename='imezemlja.xlsx')    

M=Maratonci()
M.ucitaj_maratonce()
for i in M.search_listbox('mar'):
    print(i)
M.export_maratonac(M.l_m[0])
M.export_sve_excel()
M.export_ime_zemlja_excel()